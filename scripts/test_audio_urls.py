#!/usr/bin/env python3
"""
Submit 3 audio URLs to Otto API, poll for completion, fetch full results,
and generate an Excel spreadsheet with all specified columns + extra API fields.
"""
import os
import sys
import uuid
import time
import json
from datetime import datetime

import requests
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env"))

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("openpyxl required: pip install openpyxl")
    sys.exit(1)

# ── Config ──
BASE_URL = os.getenv("OTTO_API_BASE_URL", "https://ottoai.shunyalabs.ai").rstrip("/")
API_KEY = os.getenv("OTTO_API_KEY", "5q3fwliU9ZFo3epTCsUfUiDw1Dy4DnBP")
COMPANY_ID = "1be5ea90-d3ae-4b03-8b05-f5679cd73bc4"
HEADERS = {"X-API-Key": API_KEY, "Content-Type": "application/json"}
TIMEOUT = 60
POLL_INTERVAL = 15  # seconds between status checks
MAX_POLL_TIME = 600  # max 10 minutes per call

# Add new audio URLs here before running; clear after use to avoid duplicate submissions
AUDIO_URLS = [
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/43ccc787-7f07-4a13-ac9d-d672d25a809c/4043504735.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/1fd7bea5-9ace-4e8f-a31f-152ea8269927/4015296617.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/bc946fa9-1e8a-4f1e-920c-64c4369fe778/4049722733.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/e2b0dab7-a05b-4448-8c13-fa3753f405ae/4036836500.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/6e37c8bb-16bc-4e17-867e-ae5e9f57c3b9/4037028977.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/56dc7e30-ffed-4f8d-80eb-b514ffb30a50/4050591020.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/6566e3b9-acac-4b55-aad1-5742464107fa/4058579492.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/33cc8375-04ee-43a4-8644-bdffdb8d1b1b/4060192187.mp3",
    "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/b9a3deca-d3e1-47ad-a2ad-875e58c0b7dc/4058565425.mp3",
]


def submit_call(audio_url, idx):
    """Submit a call for processing and return (call_id, job_id) or (call_id, None)."""
    call_id = f"audio_test_{idx+1}_{uuid.uuid4().hex[:8]}"
    payload = {
        "call_id": call_id,
        "company_id": COMPANY_ID,
        "audio_url": audio_url,
        "phone_number": "+14805551234",
        "rep_role": "customer_rep",
        "allow_reprocess": True,
        "metadata": {
            "agent": {
                "id": "USR_ANTHONY_ARIZONA",
                "name": "Anthony",
                "email": "anthony@arizonaroofers.com",
            }
        },
    }
    try:
        r = requests.post(
            f"{BASE_URL}/api/v1/call-processing/process",
            headers=HEADERS, json=payload, timeout=TIMEOUT,
        )
        data = r.json()
        print(f"  [{idx+1}] HTTP {r.status_code} | call_id={call_id} | job_id={data.get('job_id')}")
        if r.status_code in (202, 200):
            return call_id, data.get("job_id"), audio_url
        elif r.status_code == 409:
            # Already exists — extract job_id if available
            print(f"       409 Conflict: {data.get('detail', '')[:100]}")
            return call_id, data.get("job_id"), audio_url
        else:
            print(f"       Error: {data.get('detail', r.text[:200])}")
            return call_id, None, audio_url
    except Exception as e:
        print(f"  [{idx+1}] Exception: {e}")
        return call_id, None, audio_url


def poll_job(job_id, label=""):
    """Poll job status until completed/failed or timeout."""
    if not job_id:
        return "no_job_id"
    start = time.time()
    while time.time() - start < MAX_POLL_TIME:
        try:
            r = requests.get(
                f"{BASE_URL}/api/v1/call-processing/status/{job_id}",
                headers=HEADERS, timeout=TIMEOUT,
            )
            data = r.json()
            status = data.get("status", "unknown")
            progress = data.get("progress", {})
            pct = progress.get("percent", "?")
            step = progress.get("current_step", "?")
            print(f"    {label} status={status} progress={pct}% step={step}")
            if status in ("completed", "failed"):
                return status
        except Exception as e:
            print(f"    {label} poll error: {e}")
        time.sleep(POLL_INTERVAL)
    print(f"    {label} TIMEOUT after {MAX_POLL_TIME}s")
    return "timeout"


def fetch_summary(call_id):
    """Fetch full summary for a call."""
    try:
        r = requests.get(
            f"{BASE_URL}/api/v1/call-processing/summary/{call_id}",
            headers=HEADERS, params={"include_chunks": "true"}, timeout=TIMEOUT,
        )
        if r.status_code == 200:
            return r.json()
        print(f"    Summary fetch failed: HTTP {r.status_code}")
    except Exception as e:
        print(f"    Summary fetch error: {e}")
    return None


def fetch_detail(call_id):
    """Fetch call detail with transcript and segments."""
    try:
        r = requests.get(
            f"{BASE_URL}/api/v1/call-processing/calls/{call_id}/detail",
            headers=HEADERS,
            params={"include_transcript": "true", "include_segments": "true"},
            timeout=TIMEOUT,
        )
        if r.status_code == 200:
            return r.json()
    except Exception:
        pass
    return None


def extract_rep_name_from_transcript(segments):
    """Extract the actual CSR/rep name from transcript segments."""
    import re
    if not segments:
        return None
    intro_patterns = [
        r"(?:this is|my name is|i'm|it's|i am)\s+([A-Z][a-z]{2,15})",
        r"^(?:hey|hi|hello)[,.]?\s+.*?(?:this is|it's|i'm)\s+([A-Z][a-z]{2,15})",
    ]
    rep_segments = [s for s in segments[:15] if s.get("speaker") == "customer_rep"]
    for seg in rep_segments[:5]:
        text = seg.get("text", "")
        for pattern in intro_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                name = match.group(1).strip().title()
                false_positives = {
                    "How", "What", "The", "This", "That", "Well", "Yes", "Yeah",
                    "Okay", "Sure", "One", "Just", "Here", "There", "Let", "Can",
                    "Arizona", "Roofers", "Calling", "Thank", "Please",
                }
                if name not in false_positives and len(name) >= 3:
                    return name
    return None



def extract_row(audio_url, call_id, summary, detail):
    """Extract a flat row dict from API response data."""
    row = {}
    row["audio_url"] = audio_url
    row["Call_ID"] = call_id

    if not summary:
        return row

    # Summary section
    s = summary.get("summary", {}) or {}
    row["Summary"] = s.get("summary", "")
    row["Key Points"] = "\n".join(s.get("key_points", []))
    row["Action Items"] = "\n".join(s.get("action_items", []))
    row["Next Steps"] = "\n".join(s.get("next_steps", []))
    row["Sentiment Score"] = s.get("sentiment_score")
    row["Confidence Score (Summary)"] = s.get("confidence_score")

    # Pending actions — directly from API
    pending = s.get("pending_actions", []) or []
    row["Pending Actions Count"] = len(pending)
    if pending:
        pa_texts = []
        pa_detail_texts = []
        for i, pa in enumerate(pending):
            pa_texts.append(f"[{pa.get('type','')}] {pa.get('action_item','')} (owner: {pa.get('owner','')}, due: {pa.get('due_at','')})")
            pa_detail_texts.append(
                f"PA #{i+1}:\n"
                f"  Type: {pa.get('type','')}\n"
                f"  Action: {pa.get('action_item','')}\n"
                f"  Owner: {pa.get('owner','')}\n"
                f"  Due: {pa.get('due_at','')}\n"
                f"  Raw Text: {pa.get('raw_text','')}\n"
                f"  Confidence: {pa.get('confidence','')}\n"
                f"  Contact Method: {pa.get('contact_method','')}"
            )
        row["Pending Actions"] = "\n".join(pa_texts)
        row["Pending Actions Detail"] = "\n\n".join(pa_detail_texts)

        # Validate pending_actions against transcript
        transcript_text = ""
        if detail and detail.get("transcript"):
            transcript_text = detail["transcript"].lower()
        elif detail and detail.get("segments"):
            transcript_text = " ".join(seg.get("text", "") for seg in detail["segments"]).lower()

        if transcript_text:
            pa_validations = []
            all_pass = True
            for i, pa in enumerate(pending):
                raw_text = (pa.get("raw_text") or "").lower().strip()
                action_item = (pa.get("action_item") or "").lower().strip()
                # Check if raw_text or key phrases from action_item appear in transcript
                raw_match = False
                action_match = False
                if raw_text:
                    # Check if substantial portion of raw_text is in transcript
                    raw_words = [w for w in raw_text.split() if len(w) > 3]
                    if raw_words:
                        matched_words = sum(1 for w in raw_words if w in transcript_text)
                        raw_match = (matched_words / len(raw_words)) >= 0.5
                    else:
                        raw_match = raw_text in transcript_text
                if action_item:
                    # Check key action words against transcript
                    action_words = [w for w in action_item.split() if len(w) > 3]
                    if action_words:
                        matched_action = sum(1 for w in action_words if w in transcript_text)
                        action_match = (matched_action / len(action_words)) >= 0.4

                pa_pass = raw_match or action_match
                if not pa_pass:
                    all_pass = False
                status = "PASS" if pa_pass else "FAIL"
                pa_validations.append(
                    f"PA #{i+1} [{pa.get('type','')}]: {status}"
                    f" (raw_match={raw_match}, action_match={action_match})"
                )
            row["Pending Actions Validation"] = "\n".join(pa_validations)
            row["Pending Actions vs Transcript"] = "PASS" if all_pass else "FAIL"
        else:
            row["Pending Actions Validation"] = "No transcript available"
            row["Pending Actions vs Transcript"] = "N/A"
    else:
        row["Pending Actions"] = ""
        row["Pending Actions Detail"] = ""
        row["Pending Actions Validation"] = "No pending actions detected"
        row["Pending Actions vs Transcript"] = "N/A - No actions"

    # Qualification section
    qual = summary.get("qualification", {}) or {}
    row["Customer"] = qual.get("customer_name", "")
    row["Phone number"] = qual.get("customer_phone", "")
    row["Customer Email"] = qual.get("customer_email", "")
    row["Qualified"] = qual.get("qualification_status", "")
    row["Service Offered"] = qual.get("service_requested", "")
    row["Booked"] = qual.get("booking_status", "")
    row["Call Type"] = qual.get("detected_call_type", "")
    row["Customer_type"] = "Existing" if qual.get("is_existing_customer") else "New" if qual.get("is_existing_customer") is False else ""
    row["Call Outcome Category"] = qual.get("call_outcome_category", "")
    row["Appointment Confirmed"] = qual.get("appointment_confirmed")
    row["Appointment Date"] = qual.get("appointment_date", "")
    row["Appointment Type"] = qual.get("appointment_type", "")
    row["Appointment Timezone"] = qual.get("appointment_timezone", "")
    row["Appointment Time Confidence"] = qual.get("appointment_time_confidence")
    row["Preferred Time Window"] = qual.get("preferred_time_window", "")
    row["Appointment Intent"] = qual.get("appointment_intent", "")
    row["Service Not Offered Reason"] = qual.get("service_not_offered_reason", "")
    row["Deferred Reason"] = qual.get("deferred_reason", "")
    row["Follow Up Required"] = qual.get("follow_up_required")
    row["Follow Up Reason"] = qual.get("follow_up_reason", "")
    row["Customer Name Confidence"] = qual.get("customer_name_confidence")
    row["Address Confidence"] = qual.get("address_confidence")
    row["Confidence Score (Qual)"] = qual.get("confidence_score")

    # Scope fields (new in updated doc)
    row["Scope Classification"] = qual.get("scope_classification", "")
    row["Scope Reason"] = qual.get("scope_reason", "")
    row["Scope Confidence"] = qual.get("scope_confidence")
    row["Scope Signals"] = ", ".join(qual.get("scope_signals", []) or [])

    # BANT scores
    bant = qual.get("bant_scores", {}) or {}
    row["BANT - Budget"] = bant.get("budget")
    row["BANT - Authority"] = bant.get("authority")
    row["BANT - Need"] = bant.get("need")
    row["BANT - Timeline"] = bant.get("timeline")
    row["BANT Overall Score"] = qual.get("overall_score")

    # Decision makers
    row["Decision Makers"] = ", ".join(qual.get("decision_makers", []) or [])

    # Urgency / Budget signals
    row["Urgency Signals"] = "\n".join(qual.get("urgency_signals", []) or [])
    row["Budget Indicators"] = "\n".join(qual.get("budget_indicators", []) or [])

    # Address
    row["Service Address (Raw)"] = qual.get("service_address_raw", "")
    addr = qual.get("service_address_structured", {}) or {}
    if addr:
        row["Address Line1"] = addr.get("line1", "")
        row["Address City"] = addr.get("city", "")
        row["Address State"] = addr.get("state", "")
        row["Address Postal Code"] = addr.get("postal_code", "")
        row["Address Country"] = addr.get("country", "")

    # Property details
    prop = qual.get("property_details", {}) or {}
    if prop:
        row["Roof Type"] = prop.get("roof_type", "")
        row["Roof Age Years"] = prop.get("roof_age_years")
        row["Stories"] = prop.get("stories", "")
        row["HOA Status"] = prop.get("hoa_status", "")
        row["Has Solar"] = prop.get("has_solar")
        row["Pets"] = prop.get("pets", "")
        row["Roof Condition"] = prop.get("roof_condition", "")

    # Lead score
    lead = summary.get("lead_score", {}) or {}
    if lead:
        row["Lead Score Total"] = lead.get("total_score")
        row["Lead Band"] = lead.get("lead_band", "")
        row["Lead Confidence"] = lead.get("confidence", "")
        breakdown = lead.get("breakdown", [])
        if breakdown:
            for comp in breakdown:
                name = comp.get("component", "")
                row[f"Lead {name.title()} Points"] = f"{comp.get('points_earned',0)}/{comp.get('points_possible',0)}"
                row[f"Lead {name.title()} Reason"] = comp.get("reason", "")

    # Compliance section
    comp = summary.get("compliance", {}) or {}
    sop = comp.get("sop_compliance", {}) or {}
    row["Compliance Score"] = sop.get("score")
    row["Compliance Rate"] = sop.get("compliance_rate")
    row["Target Role"] = comp.get("target_role", "")
    row["Evaluation Mode"] = comp.get("evaluation_mode", "")
    row["SOP Version"] = sop.get("sop_version", "")

    stages = sop.get("stages", {}) or {}
    row["Stages Total"] = stages.get("total")
    row["Stages Followed"] = ", ".join(stages.get("followed", []) or [])
    row["Stages Missed"] = ", ".join(stages.get("missed", []) or [])

    row["Compliance Issues"] = "\n".join(sop.get("issues", []) or [])
    row["Positive Behaviors"] = "\n".join(sop.get("positive_behaviors", []) or [])

    # Coaching issues
    coaching_issues = sop.get("coaching_issues", []) or []
    if coaching_issues:
        ci_texts = []
        for ci in coaching_issues:
            ci_texts.append(f"[{ci.get('severity','')}] {ci.get('issue','')} | Fix: {ci.get('how_to_fix','')}")
        row["Coaching Issues"] = "\n".join(ci_texts)

    # Coaching strengths
    coaching_strengths = sop.get("coaching_strengths", []) or []
    if coaching_strengths:
        cs_texts = []
        for cs in coaching_strengths:
            cs_texts.append(f"{cs.get('behavior','')} | {cs.get('why_effective','')}")
        row["Coaching Strengths"] = "\n".join(cs_texts)

    # Objections section
    obj_data = summary.get("objections", {}) or {}
    objections = obj_data.get("objections", []) if isinstance(obj_data, dict) else []
    row["Objections Count"] = obj_data.get("total_count", len(objections))
    if objections:
        obj_texts = []
        for o in objections:
            overcome = "Resolved" if o.get("overcome") else "Unresolved"
            obj_texts.append(
                f"[{o.get('category_text','')}] \"{o.get('objection_text','')}\" "
                f"({overcome}, severity={o.get('severity','')})"
            )
        row["Objections and references"] = "\n".join(obj_texts)

    # Tags (combine key topics from summary)
    tags = set()
    if qual.get("qualification_status"):
        tags.add(qual["qualification_status"])
    if qual.get("booking_status"):
        tags.add(qual["booking_status"])
    if qual.get("detected_call_type"):
        tags.add(qual["detected_call_type"])
    if qual.get("service_requested"):
        tags.add(qual["service_requested"])
    row["Tags"] = ", ".join(tags)

    # CSR Agent — extract from transcript first, fallback to metadata
    detail_segments = detail.get("segments", []) if detail else []
    transcript_rep = extract_rep_name_from_transcript(detail_segments)
    meta = summary.get("metadata") or (detail.get("metadata") if detail else None) or {}
    agent = meta.get("agent", {}) or {}
    metadata_rep = agent.get("name") or meta.get("rep_name", "")
    row["CSR - Agent"] = transcript_rep or metadata_rep or ""

    # Call Received (call_date)
    row["Call Received"] = summary.get("processed_at") or (detail.get("call_date") if detail else "")

    # Customer Intelligence summary (combine BANT + signals into one field)
    ci_parts = []
    if bant:
        ci_parts.append(f"BANT: B={bant.get('budget','?')} A={bant.get('authority','?')} N={bant.get('need','?')} T={bant.get('timeline','?')}")
    if qual.get("overall_score"):
        ci_parts.append(f"Overall: {qual['overall_score']}")
    if qual.get("urgency_signals"):
        ci_parts.append(f"Urgency: {', '.join(qual['urgency_signals'][:2])}")
    if qual.get("budget_indicators"):
        ci_parts.append(f"Budget: {', '.join(qual['budget_indicators'][:2])}")
    row["Customer Intelligence & Lead Qualification"] = " | ".join(ci_parts)

    # Transcript
    if detail and detail.get("transcript"):
        row["Transcript"] = detail["transcript"]
    elif detail and detail.get("segments"):
        seg_texts = []
        for seg in detail["segments"]:
            seg_texts.append(f"[{seg.get('speaker','?')}] {seg.get('text','')}")
        row["Transcript"] = "\n".join(seg_texts)

    # Placeholders for manual columns
    row["Old data displayed on dashboard"] = ""
    row["Point raised by Tushar on mail"] = ""
    row["QC Status"] = ""
    row["Reason"] = ""
    row["Comments"] = ""

    return row


def write_excel(rows, output_path):
    """Write rows to Excel with formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audio URL Test Results"

    # Ordered columns matching user's requested order, then extras
    primary_cols = [
        "Call Received", "audio_url", "CSR - Agent", "Customer", "Phone number",
        "Qualified", "Service Offered", "Booked",
        "Customer Intelligence & Lead Qualification",
        "Objections and references",
        "Pending Actions Count", "Pending Actions", "Pending Actions Detail",
        "Pending Actions vs Transcript", "Pending Actions Validation",
        "Tags", "Call Type", "Customer_type",
        "Old data displayed on dashboard", "Transcript",
        "Point raised by Tushar on mail", "QC Status", "Reason",
        "Call_ID", "Comments",
    ]

    # Collect all extra columns from API data
    all_keys = set()
    for r in rows:
        all_keys.update(r.keys())
    extra_cols = sorted(k for k in all_keys if k not in primary_cols)

    all_cols = primary_cols + extra_cols

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, col_name in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(all_cols, 1):
            val = row_data.get(col_name, "")
            if val is None:
                val = ""
            elif isinstance(val, bool):
                val = "Yes" if val else "No"
            elif isinstance(val, float):
                val = round(val, 3)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

    # Auto-width (capped)
    for col_idx, col_name in enumerate(all_cols, 1):
        max_len = len(str(col_name))
        for row_idx in range(2, len(rows) + 2):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            first_line = val.split("\n")[0] if val else ""
            max_len = max(max_len, min(len(first_line), 60))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 3, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add an index sheet mapping audio URL -> call_id
    ws2 = wb.create_sheet("Audio URL Mapping")
    ws2.cell(row=1, column=1, value="Audio URL #").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="Audio URL").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=3, value="Call ID").font = header_font
    ws2.cell(row=1, column=3).fill = header_fill
    ws2.cell(row=1, column=4, value="Status").font = header_font
    ws2.cell(row=1, column=4).fill = header_fill
    for i, r in enumerate(rows):
        ws2.cell(row=i+2, column=1, value=i+1)
        ws2.cell(row=i+2, column=2, value=r.get("audio_url", ""))
        ws2.cell(row=i+2, column=3, value=r.get("Call_ID", ""))
        ws2.cell(row=i+2, column=4, value="Completed" if r.get("Customer") else "No Data")
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 80
    ws2.column_dimensions["C"].width = 40
    ws2.column_dimensions["D"].width = 15

    wb.save(output_path)
    print(f"\nExcel saved: {output_path}")


def main():
    output_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "reports",
        "audio_url_test_results_final.xlsx",
    )
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    print(f"Otto Audio URL Test Script")
    print(f"API: {BASE_URL}")
    print(f"Company: {COMPANY_ID}")
    print(f"Audio URLs: {len(AUDIO_URLS)}")
    print(f"Output: {output_path}")
    print("=" * 60)

    # Step 1: Submit all calls
    print("\n[1/4] Submitting calls...")
    submissions = []
    for i, url in enumerate(AUDIO_URLS):
        call_id, job_id, audio_url = submit_call(url, i)
        submissions.append((call_id, job_id, audio_url))

    # Step 2: Poll for completion
    print("\n[2/4] Polling for completion...")
    statuses = {}
    for call_id, job_id, audio_url in submissions:
        label = f"Call {call_id[:30]}"
        if job_id:
            statuses[call_id] = poll_job(job_id, label)
        else:
            statuses[call_id] = "submit_failed"

    # Step 3: Fetch results
    print("\n[3/4] Fetching summaries and details...")
    rows = []
    for call_id, job_id, audio_url in submissions:
        print(f"  Fetching {call_id}...")
        summary = fetch_summary(call_id) if statuses[call_id] == "completed" else None
        detail = fetch_detail(call_id) if statuses[call_id] == "completed" else None
        row = extract_row(audio_url, call_id, summary, detail)
        row["Processing Status"] = statuses[call_id]
        rows.append(row)

    # Step 4: Write Excel
    print("\n[4/4] Writing Excel spreadsheet...")
    write_excel(rows, output_path)

    # Print mapping summary
    print("\n" + "=" * 60)
    print("AUDIO URL -> CALL ID MAPPING:")
    print("-" * 60)
    completed_ids = []
    for i, (call_id, job_id, audio_url) in enumerate(submissions):
        short_url = audio_url.split("/")[-1]
        status = statuses[call_id]
        cust = rows[i].get("Customer", "N/A")
        print(f"  URL #{i+1}: {short_url}")
        print(f"    call_id: {call_id}")
        print(f"    status:  {status}")
        print(f"    customer: {cust}")
        print()
        if status == "completed":
            completed_ids.append(call_id)

    # Step 5: Generate HTML report via generate_new_report.py
    if completed_ids:
        import subprocess
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        report_script = os.path.join(project_root, "generate_new_report.py")
        html_output = os.path.join(project_root, "reports", "audio_url_test_report.html")
        call_ids_str = ",".join(completed_ids)
        print(f"\n[5/5] Generating HTML report for {len(completed_ids)} completed calls...")
        subprocess.run(
            [sys.executable, report_script, "--call-ids", call_ids_str, "--output", html_output],
            cwd=project_root,
        )


if __name__ == "__main__":
    main()
