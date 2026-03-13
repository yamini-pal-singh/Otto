#!/usr/bin/env python3
"""
CSR Test Runner — Tests only the 10 specific CSR audio calls,
updates the existing Excel sheet, and generates an HTML report.
"""
import os
import json
import html as html_mod
from datetime import datetime
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from copy import copy

# ─────────────────────────────────────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────────────────────────────────────
BASE_URL = "https://ottoai.shunyalabs.ai"
API_KEY  = "5q3fwliU9ZFo3epTCsUfUiDw1Dy4DnBP"
COMPANY_ID = "1be5ea90-d3ae-4b03-8b05-f5679cd73bc4"
HEADERS = {"X-API-Key": API_KEY, "Content-Type": "application/json"}
TIMEOUT = 60

EXCEL_PATH = "/Users/unitedwecare/Otto_repos/reports/audio_url_test_results_final.xlsx"
HTML_OUTPUT = "/Users/unitedwecare/Otto_repos/call_report_csr_10.html"

# The 10 CSR audio URLs and their known call IDs — March 13, 2026 batch
CSR_CALLS = [
    {"num": 1,  "audio_url": "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/bd24a254-fe4e-48ef-9c3e-3c3f3adfd68c/4082148782.mp3",  "call_id": "scenario_test_1_88148c4c"},
    {"num": 2,  "audio_url": "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/00273c23-da2c-4e1b-b49c-4ff85d4a766a/4082112377.mp3",  "call_id": "scenario_test_2_2e229983"},
    {"num": 3,  "audio_url": "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/c0183543-60e4-434f-a247-89ec9ef8e1e3/4081871096.mp3",  "call_id": "scenario_test_3_cc5d0fb9"},
    {"num": 4,  "audio_url": "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/eb359c0a-02e2-4e94-b018-b3585c8a4024/4081765241.mp3",  "call_id": "scenario_test_4_334fff68"},
    {"num": 5,  "audio_url": "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/4da88a74-f129-481d-92d4-626c5a728835/4081742162.mp3",  "call_id": "scenario_test_5_8b5c1e22"},
    {"num": 6,  "audio_url": "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/c6c4fff3-72d5-42d5-9cb4-16b8272e3d3b/4079556188.mp3",  "call_id": "scenario_test_6_c383236d"},
    {"num": 7,  "audio_url": "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/ddccb9c0-fce8-4eb5-81bd-5e738cb95979/4078581674.mp3",  "call_id": "scenario_test_7_b228486b"},
    {"num": 8,  "audio_url": "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/add722b1-c29f-4615-b389-4d29773d5958/st_309419607.mp3",  "call_id": "scenario_test_8_a99fa915"},
    {"num": 9,  "audio_url": "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/df6c19b0-5d39-457b-8053-4dd83f4407cf/4076546354.mp3",  "call_id": "scenario_test_9_e37ecbc8"},
    {"num": 10, "audio_url": "https://ottoaudio.s3.ap-southeast-2.amazonaws.com/recordings/6dd70da4-9d39-4496-871e-827af1429e29/4077402764.mp3",  "call_id": "scenario_test_10_fcee5ce4"},
]


# ─────────────────────────────────────────────────────────────────────────────
# API Helpers
# ─────────────────────────────────────────────────────────────────────────────
def fetch_summary(call_id):
    try:
        r = requests.get(f"{BASE_URL}/api/v1/call-processing/summary/{call_id}",
                         headers=HEADERS, timeout=TIMEOUT)
        if r.status_code == 200:
            return r.json()
    except Exception as e:
        print(f"    ERROR fetching summary for {call_id}: {e}")
    return None

def fetch_detail(call_id):
    try:
        r = requests.get(f"{BASE_URL}/api/v1/call-processing/calls/{call_id}/detail",
                         headers=HEADERS,
                         params={"include_transcript": "true", "include_segments": "true"},
                         timeout=TIMEOUT)
        if r.status_code == 200:
            return r.json()
    except Exception as e:
        print(f"    ERROR fetching detail for {call_id}: {e}")
    return None

def find_call_by_audio_url(audio_url):
    """Search all calls to find one matching this audio URL."""
    try:
        r = requests.get(f"{BASE_URL}/api/v1/call-processing/calls",
                         headers=HEADERS,
                         params={"company_id": COMPANY_ID, "limit": 50, "sort_by": "call_date", "sort_order": "desc"},
                         timeout=TIMEOUT)
        if r.status_code == 200:
            calls = r.json().get("calls", [])
            for c in calls:
                if c.get("audio_url") == audio_url:
                    return c.get("call_id")
    except Exception as e:
        print(f"    ERROR searching for audio URL: {e}")
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Data Extraction
# ─────────────────────────────────────────────────────────────────────────────
def extract_transcript_text(detail):
    """Extract full transcript as text from detail segments."""
    if not detail or not isinstance(detail, dict):
        return ""
    transcript_obj = detail.get("transcript", {})
    if isinstance(transcript_obj, dict):
        segments = transcript_obj.get("segments", []) or []
    else:
        segments = []
    if not segments:
        segments = detail.get("segments", []) or []
    lines = []
    for seg in segments:
        speaker = seg.get("speaker", "unknown")
        text = seg.get("text", "")
        lines.append(f"{speaker}: {text}")
    return "\n".join(lines)

def extract_rep_name(detail):
    """Extract CSR name from transcript intro."""
    import re
    if not detail or not isinstance(detail, dict):
        return None
    transcript_obj = detail.get("transcript", {})
    if isinstance(transcript_obj, dict):
        segments = transcript_obj.get("segments", []) or []
    else:
        segments = []
    if not segments:
        segments = detail.get("segments", []) or []
    intro_patterns = [
        r"(?:this is|my name is|i'm|it's|i am)\s+([A-Z][a-z]{2,15})",
        r"^(?:hey|hi|hello)[,.]?\s+.*?(?:this is|it's|i'm)\s+([A-Z][a-z]{2,15})",
    ]
    rep_segs = [s for s in segments[:15] if s.get("speaker") == "customer_rep"]
    for seg in rep_segs[:5]:
        text = seg.get("text", "")
        for pat in intro_patterns:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                name = m.group(1).strip().title()
                if name not in {"How","What","The","This","That","Well","Yes","Yeah",
                                "Okay","Sure","One","Just","Here","There","Let","Can",
                                "Arizona","Roofers","Calling","Thank","Please"} and len(name) >= 3:
                    return name
    return None

def safe_join(lst, sep=", "):
    if not lst:
        return ""
    if isinstance(lst, list):
        return sep.join(str(x) for x in lst)
    return str(lst)

def extract_row_data(call_entry, summary, detail):
    """Extract all fields for the Excel row."""
    s = summary or {}
    sum_obj = s.get("summary", {}) or {}
    if isinstance(sum_obj, str):
        sum_obj = {}
    qual = s.get("qualification", {}) or {}
    comp = s.get("compliance", {}) or {}
    sop  = comp.get("sop_compliance", {}) or {}
    stages = sop.get("stages", {}) or {}
    bant = qual.get("bant_scores", {}) or {}
    obj_data = s.get("objections", {}) or {}
    objections = obj_data.get("objections", []) or []
    pending_actions = sum_obj.get("pending_actions", []) or []
    addr = qual.get("service_address_structured", {}) or {}
    cust_intel = s.get("customer_intelligence", {}) or {}

    rep_name = extract_rep_name(detail)
    transcript = extract_transcript_text(detail)

    # Pending actions
    pa_count = len(pending_actions)
    pa_types = safe_join([pa.get("type","") for pa in pending_actions])
    pa_detail = safe_join([pa.get("action_item","") or pa.get("raw_text","") for pa in pending_actions], " | ")

    # Compliance issues
    coaching_issues = sop.get("coaching_issues", []) or []
    compliance_issues = safe_join([c.get("issue","") if isinstance(c,dict) else str(c) for c in coaching_issues])
    coaching_strengths = safe_join(sop.get("positive_behaviors", []) or [])

    # Stages
    followed = stages.get("followed", []) or []
    missed = stages.get("missed", []) or []
    total_stages = len(followed) + len(missed)

    # Objections
    obj_count = len(objections)
    obj_text = safe_join([o.get("objection_text","") for o in objections], " | ")

    return {
        "call_received": qual.get("call_date") or detail.get("call_date", "") if detail else "",
        "csr_agent": rep_name or "",
        "customer": qual.get("customer_name") or "",
        "phone": qual.get("phone_number") or call_entry.get("phone_number", ""),
        "qualified": qual.get("qualification_status") or "",
        "service_offered": qual.get("service_requested") or "",
        "booked": qual.get("booking_status") or "",
        "cust_intel": str(cust_intel)[:500] if cust_intel else "",
        "objections_refs": obj_text,
        "pa_count": pa_count,
        "pa_types": pa_types,
        "pa_detail": pa_detail,
        "pa_vs_transcript": "",
        "pa_validation": "",
        "tags": "",
        "call_type": qual.get("detected_call_type") or "",
        "customer_type": "existing" if qual.get("is_existing_customer") else "new",
        "old_data": "",
        "transcript": transcript[:5000],
        "tushar_point": "",
        "qc_status": "",
        "reason": "",
        "action_items": safe_join(sum_obj.get("action_items", []) or []),
        "addr_city": addr.get("city") or "",
        "addr_confidence": qual.get("address_confidence") or "",
        "addr_country": addr.get("country") or "",
        "addr_line1": addr.get("line1") or "",
        "addr_postal": addr.get("postal_code") or "",
        "addr_state": addr.get("state") or "",
        "appt_confirmed": qual.get("appointment_confirmed"),
        "appt_date": qual.get("appointment_date") or "",
        "appt_intent": qual.get("appointment_intent") or "",
        "appt_time_confidence": qual.get("appointment_time_confidence") or "",
        "appt_timezone": qual.get("appointment_timezone") or "",
        "appt_type": qual.get("appointment_type") or "",
        "bant_authority": bant.get("authority"),
        "bant_budget": bant.get("budget"),
        "bant_need": bant.get("need"),
        "bant_timeline": bant.get("timeline"),
        "bant_overall": qual.get("overall_score"),
        "budget_indicators": safe_join(qual.get("budget_indicators") or []),
        "call_outcome": qual.get("call_outcome_category") or "",
        "coaching_issues": compliance_issues,
        "coaching_strengths": coaching_strengths,
        "compliance_issues_detail": safe_join([c.get("description","") if isinstance(c,dict) else str(c) for c in coaching_issues]),
        "compliance_rate": f"{int(float(sop.get('score',0) or 0)*100)}%" if sop.get("score") else "",
        "compliance_score": sop.get("score"),
        "confidence_qual": qual.get("confidence_score"),
        "confidence_summary": sum_obj.get("confidence_score"),
        "customer_email": qual.get("customer_email") or "",
        "customer_name_confidence": qual.get("customer_name_confidence"),
        "decision_makers": safe_join(qual.get("decision_makers") or []),
        "deferred_reason": qual.get("deferred_reason") or "",
        "evaluation_mode": s.get("evaluation_mode") or "",
        "follow_up_reason": qual.get("follow_up_reason") or "",
        "follow_up_required": qual.get("follow_up_required"),
        "key_points": safe_join(sum_obj.get("key_points", []) or []),
        "next_steps": safe_join(sum_obj.get("next_steps", []) or []),
        "objections_count": obj_count,
        "positive_behaviors": coaching_strengths,
        "preferred_time": qual.get("preferred_time_window") or "",
        "processing_status": s.get("processing_status") or "completed",
        "sop_version": s.get("sop_version") or "",
        "scope_classification": qual.get("scope_classification") or "",
        "scope_confidence": qual.get("scope_confidence") or "",
        "scope_reason": qual.get("scope_reason") or "",
        "scope_signals": safe_join(qual.get("scope_signals") or []),
        "sentiment": sum_obj.get("sentiment_score"),
        "service_addr_raw": qual.get("service_address_raw") or "",
        "service_not_offered_reason": qual.get("service_not_offered_reason") or "",
        "stages_followed": safe_join(followed),
        "stages_missed": safe_join(missed),
        "stages_total": total_stages,
        "summary_text": sum_obj.get("summary") or "",
        "target_role": qual.get("target_role") or "",
        "urgency_signals": safe_join(qual.get("urgency_signals") or []),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Excel Updater
# ─────────────────────────────────────────────────────────────────────────────
def update_excel(results):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Audio URL Test Results"]
    ws_map = wb["Audio URL Mapping"]

    # Map audio_url -> row number in results sheet (row index 2 = first data row)
    audio_to_row = {}
    for row_idx in range(2, ws.max_row + 1):
        cell_audio = ws.cell(row_idx, 2).value  # col B = audio_url
        if cell_audio:
            audio_to_row[cell_audio.strip()] = row_idx

    # Column mapping: header -> col index
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i+1 for i, h in enumerate(headers) if h}

    def set_cell(row, header, value):
        c = col.get(header)
        if c and value is not None and value != "":
            ws.cell(row, c).value = value

    updated_count = 0
    new_rows = []

    for entry in results:
        audio_url = entry["audio_url"]
        data = entry["data"]
        call_id = entry["call_id"]
        status = entry["status"]

        if audio_url in audio_to_row:
            row_idx = audio_to_row[audio_url]
        else:
            # New row — append
            row_idx = ws.max_row + 1
            ws.cell(row_idx, col.get("audio_url", 2)).value = audio_url
            new_rows.append(row_idx)

        ws.cell(row_idx, col.get("Call_ID", 24)).value = call_id

        if data:
            set_cell(row_idx, "Call Received",             data["call_received"])
            set_cell(row_idx, "CSR - Agent",               data["csr_agent"])
            set_cell(row_idx, "Customer",                  data["customer"])
            set_cell(row_idx, "Phone number",              data["phone"])
            set_cell(row_idx, "Qualified",                 data["qualified"])
            set_cell(row_idx, "Service Offered",           data["service_offered"])
            set_cell(row_idx, "Booked",                    data["booked"])
            set_cell(row_idx, "Customer Intelligence & Lead Qualification", data["cust_intel"])
            set_cell(row_idx, "Objections and references", data["objections_refs"])
            set_cell(row_idx, "Pending Actions Count",     data["pa_count"])
            set_cell(row_idx, "Pending Actions",           data["pa_types"])
            set_cell(row_idx, "Pending Actions Detail",    data["pa_detail"])
            set_cell(row_idx, "Call Type",                 data["call_type"])
            set_cell(row_idx, "Customer_type",             data["customer_type"])
            set_cell(row_idx, "Transcript",                data["transcript"])
            set_cell(row_idx, "Action Items",              data["action_items"])
            set_cell(row_idx, "Address City",              data["addr_city"])
            set_cell(row_idx, "Address Confidence",        data["addr_confidence"])
            set_cell(row_idx, "Address Country",           data["addr_country"])
            set_cell(row_idx, "Address Line1",             data["addr_line1"])
            set_cell(row_idx, "Address Postal Code",       data["addr_postal"])
            set_cell(row_idx, "Address State",             data["addr_state"])
            set_cell(row_idx, "Appointment Confirmed",     data["appt_confirmed"])
            set_cell(row_idx, "Appointment Date",          data["appt_date"])
            set_cell(row_idx, "Appointment Intent",        data["appt_intent"])
            set_cell(row_idx, "Appointment Time Confidence", data["appt_time_confidence"])
            set_cell(row_idx, "Appointment Timezone",      data["appt_timezone"])
            set_cell(row_idx, "Appointment Type",          data["appt_type"])
            set_cell(row_idx, "BANT - Authority",          data["bant_authority"])
            set_cell(row_idx, "BANT - Budget",             data["bant_budget"])
            set_cell(row_idx, "BANT - Need",               data["bant_need"])
            set_cell(row_idx, "BANT - Timeline",           data["bant_timeline"])
            set_cell(row_idx, "BANT Overall Score",        data["bant_overall"])
            set_cell(row_idx, "Budget Indicators",         data["budget_indicators"])
            set_cell(row_idx, "Call Outcome Category",     data["call_outcome"])
            set_cell(row_idx, "Coaching Issues",           data["coaching_issues"])
            set_cell(row_idx, "Coaching Strengths",        data["coaching_strengths"])
            set_cell(row_idx, "Compliance Issues",         data["compliance_issues_detail"])
            set_cell(row_idx, "Compliance Rate",           data["compliance_rate"])
            set_cell(row_idx, "Compliance Score",          data["compliance_score"])
            set_cell(row_idx, "Confidence Score (Qual)",   data["confidence_qual"])
            set_cell(row_idx, "Confidence Score (Summary)",data["confidence_summary"])
            set_cell(row_idx, "Customer Email",            data["customer_email"])
            set_cell(row_idx, "Customer Name Confidence",  data["customer_name_confidence"])
            set_cell(row_idx, "Decision Makers",           data["decision_makers"])
            set_cell(row_idx, "Deferred Reason",           data["deferred_reason"])
            set_cell(row_idx, "Evaluation Mode",           data["evaluation_mode"])
            set_cell(row_idx, "Follow Up Reason",          data["follow_up_reason"])
            set_cell(row_idx, "Follow Up Required",        data["follow_up_required"])
            set_cell(row_idx, "Key Points",                data["key_points"])
            set_cell(row_idx, "Next Steps",                data["next_steps"])
            set_cell(row_idx, "Objections Count",          data["objections_count"])
            set_cell(row_idx, "Positive Behaviors",        data["positive_behaviors"])
            set_cell(row_idx, "Preferred Time Window",     data["preferred_time"])
            set_cell(row_idx, "Processing Status",         data["processing_status"])
            set_cell(row_idx, "SOP Version",               data["sop_version"])
            set_cell(row_idx, "Scope Classification",      data["scope_classification"])
            set_cell(row_idx, "Scope Confidence",          data["scope_confidence"])
            set_cell(row_idx, "Scope Reason",              data["scope_reason"])
            set_cell(row_idx, "Scope Signals",             data["scope_signals"])
            set_cell(row_idx, "Sentiment Score",           data["sentiment"])
            set_cell(row_idx, "Service Address (Raw)",     data["service_addr_raw"])
            set_cell(row_idx, "Service Not Offered Reason",data["service_not_offered_reason"])
            set_cell(row_idx, "Stages Followed",           data["stages_followed"])
            set_cell(row_idx, "Stages Missed",             data["stages_missed"])
            set_cell(row_idx, "Stages Total",              data["stages_total"])
            set_cell(row_idx, "Summary",                   data["summary_text"])
            set_cell(row_idx, "Target Role",               data["target_role"])
            set_cell(row_idx, "Urgency Signals",           data["urgency_signals"])
            updated_count += 1

    # Update Audio URL Mapping sheet
    map_rows = {ws_map.cell(r, 2).value: r for r in range(2, ws_map.max_row + 1) if ws_map.cell(r, 2).value}
    for entry in results:
        audio_url = entry["audio_url"]
        call_id = entry["call_id"]
        num = entry["num"]
        status_str = "Completed" if entry["data"] else "No Data"
        if audio_url in map_rows:
            r = map_rows[audio_url]
            ws_map.cell(r, 3).value = call_id
            ws_map.cell(r, 4).value = status_str
        else:
            new_r = ws_map.max_row + 1
            ws_map.cell(new_r, 1).value = num
            ws_map.cell(new_r, 2).value = audio_url
            ws_map.cell(new_r, 3).value = call_id
            ws_map.cell(new_r, 4).value = status_str

    wb.save(EXCEL_PATH)
    print(f"  Excel updated: {updated_count} rows with data, {len(new_rows)} new rows added")
    return updated_count


# ─────────────────────────────────────────────────────────────────────────────
# HTML Report Builder
# ─────────────────────────────────────────────────────────────────────────────
def esc(t):
    return html_mod.escape(str(t)) if t is not None else ""

def score_bar(score, width=80):
    if score is None:
        return '<span style="color:#94a3b8">N/A</span>'
    pct = int(float(score) * 100) if float(score) <= 1 else int(float(score))
    color = "#22c55e" if pct >= 80 else ("#eab308" if pct >= 60 else ("#f97316" if pct >= 40 else "#ef4444"))
    return (f'<div style="display:flex;align-items:center;gap:6px">'
            f'<div style="background:#334155;border-radius:4px;height:8px;width:{width}px">'
            f'<div style="background:{color};height:100%;width:{pct}%;border-radius:4px"></div></div>'
            f'<span style="color:{color};font-weight:700">{pct}%</span></div>')

def bool_badge(val):
    if val is True or str(val).lower() in ("true","yes","1"):
        return '<span style="background:#22c55e;color:#fff;padding:2px 8px;border-radius:9999px;font-size:11px">YES</span>'
    if val is False or str(val).lower() in ("false","no","0"):
        return '<span style="background:#ef4444;color:#fff;padding:2px 8px;border-radius:9999px;font-size:11px">NO</span>'
    return f'<span style="color:#94a3b8">{esc(val)}</span>'

def status_badge(s):
    color = {"completed":"#22c55e","processing":"#3b82f6","queued":"#eab308","failed":"#ef4444","no data":"#94a3b8"}.get(str(s).lower(),"#94a3b8")
    return f'<span style="background:{color};color:#fff;padding:2px 8px;border-radius:9999px;font-size:11px;font-weight:600">{esc(s).upper()}</span>'

def pa_badge(count):
    if count == 0:
        return '<span style="background:#22c55e;color:#fff;padding:2px 8px;border-radius:9999px;font-size:12px;font-weight:700">0</span>'
    return f'<span style="background:#f97316;color:#fff;padding:2px 8px;border-radius:9999px;font-size:12px;font-weight:700">{count}</span>'

def build_html(results):
    now = datetime.now().strftime("%B %d, %Y at %I:%M %p")
    total = len(results)
    with_data = sum(1 for r in results if r["data"])
    no_data = total - with_data
    booked = sum(1 for r in results if r["data"] and str(r["data"].get("booked","")).lower() in ("booked","confirmed","yes","true"))
    total_pa = sum(r["data"].get("pa_count",0) for r in results if r["data"])

    # Build cards
    cards_html = ""
    for entry in results:
        num = entry["num"]
        call_id = entry["call_id"]
        audio_url = entry["audio_url"]
        d = entry["data"]
        status = "Completed" if d else "No Data"

        if not d:
            cards_html += f'''
            <div class="call-card no-data">
                <div class="call-header">
                    <span class="call-num">Call #{num}</span>
                    {status_badge("no data")}
                    <span style="color:#64748b;font-size:12px;margin-left:auto">{esc(call_id)}</span>
                </div>
                <p style="color:#64748b;padding:16px">No data available for this call. The audio may still be processing.</p>
                <div style="padding:8px 16px;color:#475569;font-size:12px">Audio: {esc(audio_url[:70])}...</div>
            </div>'''
            continue

        rep = d.get("csr_agent") or "Unknown CSR"
        customer = d.get("customer") or "Unknown"
        call_type = d.get("call_type") or "N/A"
        qualified = d.get("qualified") or "N/A"
        booked_val = d.get("booked") or "N/A"
        pa_count = d.get("pa_count", 0)
        compliance = d.get("compliance_score")
        sentiment = d.get("sentiment")
        summary_text = d.get("summary_text") or ""
        key_points = d.get("key_points") or ""
        action_items = d.get("action_items") or ""
        next_steps = d.get("next_steps") or ""
        pa_detail = d.get("pa_detail") or ""
        coaching = d.get("coaching_issues") or ""
        missed_stages = d.get("stages_missed") or ""
        followed_stages = d.get("stages_followed") or ""
        addr_raw = d.get("service_addr_raw") or ""
        appt_date = d.get("appt_date") or ""
        appt_confirmed = d.get("appt_confirmed")
        bant_budget = d.get("bant_budget")
        bant_need = d.get("bant_need")
        bant_authority = d.get("bant_authority")
        bant_timeline = d.get("bant_timeline")
        bant_overall = d.get("bant_overall")
        objections_text = d.get("objections_refs") or ""
        urgency = d.get("urgency_signals") or ""
        transcript = d.get("transcript") or ""
        call_outcome = d.get("call_outcome") or ""
        cust_type = d.get("customer_type") or ""
        follow_up = d.get("follow_up_required")
        follow_up_reason = d.get("follow_up_reason") or ""

        cards_html += f'''
        <div class="call-card">
            <div class="call-header">
                <span class="call-num">Call #{num}</span>
                {status_badge("completed")}
                <span class="rep-name">CSR: {esc(rep)}</span>
                <span class="customer-name">Customer: {esc(customer)}</span>
                <span style="color:#64748b;font-size:11px;margin-left:auto">{esc(call_id)}</span>
            </div>

            <div class="call-meta-row">
                <div class="meta-item"><span class="meta-label">Call Type</span><span class="meta-val">{esc(call_type)}</span></div>
                <div class="meta-item"><span class="meta-label">Customer Type</span><span class="meta-val">{esc(cust_type)}</span></div>
                <div class="meta-item"><span class="meta-label">Qualified</span><span class="meta-val">{esc(qualified)}</span></div>
                <div class="meta-item"><span class="meta-label">Booked</span><span class="meta-val">{esc(booked_val)}</span></div>
                <div class="meta-item"><span class="meta-label">Outcome</span><span class="meta-val">{esc(call_outcome)}</span></div>
                <div class="meta-item"><span class="meta-label">Pending Actions</span><span class="meta-val">{pa_badge(pa_count)}</span></div>
                <div class="meta-item"><span class="meta-label">Appt Confirmed</span><span class="meta-val">{bool_badge(appt_confirmed)}</span></div>
                <div class="meta-item"><span class="meta-label">Follow-Up</span><span class="meta-val">{bool_badge(follow_up)}</span></div>
            </div>

            <div class="scores-row">
                <div class="score-box"><div class="score-label">Compliance</div>{score_bar(compliance)}</div>
                <div class="score-box"><div class="score-label">Sentiment</div>{score_bar(sentiment)}</div>
                <div class="score-box"><div class="score-label">BANT Overall</div>{score_bar(bant_overall)}</div>
                <div class="score-box bant-detail">
                    <div class="score-label">BANT Breakdown</div>
                    <div class="bant-grid">
                        <span>Budget: <b>{f"{int(float(bant_budget)*100)}%" if bant_budget is not None else "N/A"}</b></span>
                        <span>Authority: <b>{f"{int(float(bant_authority)*100)}%" if bant_authority is not None else "N/A"}</b></span>
                        <span>Need: <b>{f"{int(float(bant_need)*100)}%" if bant_need is not None else "N/A"}</b></span>
                        <span>Timeline: <b>{f"{int(float(bant_timeline)*100)}%" if bant_timeline is not None else "N/A"}</b></span>
                    </div>
                </div>
            </div>

            <div class="sections-grid">
                <div class="section-box">
                    <div class="section-title">Summary</div>
                    <p>{esc(summary_text)}</p>
                </div>
                {"" if not pa_detail else f'<div class="section-box pa-box"><div class="section-title">Pending Actions ({pa_count})</div><p>{esc(pa_detail)}</p></div>'}
                {"" if not key_points else f'<div class="section-box"><div class="section-title">Key Points</div><p>{esc(key_points)}</p></div>'}
                {"" if not action_items else f'<div class="section-box"><div class="section-title">Action Items</div><p>{esc(action_items)}</p></div>'}
                {"" if not next_steps else f'<div class="section-box"><div class="section-title">Next Steps</div><p>{esc(next_steps)}</p></div>'}
                {"" if not coaching else f'<div class="section-box warn-box"><div class="section-title">Coaching Issues</div><p>{esc(coaching)}</p></div>'}
                {"" if not missed_stages else f'<div class="section-box warn-box"><div class="section-title">Missed Stages</div><p>{esc(missed_stages)}</p></div>'}
                {"" if not objections_text else f'<div class="section-box"><div class="section-title">Objections</div><p>{esc(objections_text)}</p></div>'}
                {"" if not urgency else f'<div class="section-box"><div class="section-title">Urgency Signals</div><p>{esc(urgency)}</p></div>'}
                {"" if not addr_raw else f'<div class="section-box"><div class="section-title">Service Address</div><p>{esc(addr_raw)}</p></div>'}
                {"" if not appt_date else f'<div class="section-box"><div class="section-title">Appointment</div><p>{esc(appt_date)}</p></div>'}
                {"" if not follow_up_reason else f'<div class="section-box"><div class="section-title">Follow-Up Reason</div><p>{esc(follow_up_reason)}</p></div>'}
            </div>

            {("" if not transcript else
              '<details class="transcript-toggle"><summary>View Transcript</summary><pre class="transcript-text">' + esc(transcript) + '</pre></details>')}
        </div>'''

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>CSR Test Report — 10 Calls | Arizona Roofers</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:#0f172a;color:#e2e8f0;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;font-size:14px;line-height:1.5}}
  .page-header{{background:linear-gradient(135deg,#1e293b 0%,#0f172a 100%);padding:32px 40px;border-bottom:1px solid #1e293b}}
  .page-header h1{{font-size:26px;font-weight:700;color:#f1f5f9}}
  .page-header .sub{{color:#64748b;margin-top:4px}}
  .summary-bar{{display:flex;gap:16px;padding:20px 40px;background:#1e293b;border-bottom:1px solid #334155;flex-wrap:wrap}}
  .stat-card{{background:#0f172a;border:1px solid #334155;border-radius:10px;padding:12px 20px;min-width:110px;text-align:center}}
  .stat-val{{font-size:28px;font-weight:800;display:block}}
  .stat-label{{font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:.05em}}
  .calls-container{{padding:24px 40px;display:flex;flex-direction:column;gap:20px}}
  .call-card{{background:#1e293b;border:1px solid #334155;border-radius:12px;overflow:hidden}}
  .call-card.no-data{{opacity:.6}}
  .call-header{{display:flex;align-items:center;gap:12px;padding:14px 20px;background:#1a2744;border-bottom:1px solid #334155;flex-wrap:wrap}}
  .call-num{{font-size:16px;font-weight:800;color:#60a5fa}}
  .rep-name{{font-size:13px;color:#94a3b8;background:#1e293b;padding:2px 10px;border-radius:6px}}
  .customer-name{{font-size:13px;color:#a5b4fc;background:#1e293b;padding:2px 10px;border-radius:6px}}
  .call-meta-row{{display:flex;flex-wrap:wrap;gap:0;border-bottom:1px solid #334155}}
  .meta-item{{padding:10px 16px;border-right:1px solid #334155;flex:1;min-width:120px}}
  .meta-label{{display:block;font-size:10px;color:#64748b;text-transform:uppercase;letter-spacing:.05em;margin-bottom:3px}}
  .meta-val{{font-size:13px;font-weight:600;color:#cbd5e1}}
  .scores-row{{display:flex;flex-wrap:wrap;gap:0;padding:12px 16px;background:#162032;border-bottom:1px solid #334155}}
  .score-box{{padding:8px 16px;flex:1;min-width:160px}}
  .score-label{{font-size:10px;color:#64748b;text-transform:uppercase;margin-bottom:6px}}
  .bant-detail{{min-width:280px}}
  .bant-grid{{display:grid;grid-template-columns:1fr 1fr;gap:4px 16px;font-size:12px;color:#94a3b8}}
  .bant-grid b{{color:#e2e8f0}}
  .sections-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:12px;padding:16px}}
  .section-box{{background:#162032;border:1px solid #334155;border-radius:8px;padding:12px 14px}}
  .section-box.pa-box{{border-color:#f97316;background:#1f1208}}
  .section-box.warn-box{{border-color:#eab308;background:#1a1600}}
  .section-title{{font-size:10px;text-transform:uppercase;letter-spacing:.06em;color:#64748b;margin-bottom:6px;font-weight:700}}
  .section-box p{{color:#cbd5e1;font-size:13px;line-height:1.5}}
  .transcript-toggle{{border-top:1px solid #334155}}
  .transcript-toggle summary{{padding:10px 20px;cursor:pointer;color:#64748b;font-size:12px;user-select:none}}
  .transcript-toggle summary:hover{{color:#94a3b8}}
  .transcript-text{{background:#0f172a;padding:16px 20px;font-size:12px;line-height:1.7;color:#94a3b8;white-space:pre-wrap;word-break:break-word;max-height:400px;overflow-y:auto}}
</style>
</head>
<body>
<div class="page-header">
  <h1>CSR Test Report — 10 Calls</h1>
  <div class="sub">Arizona Roofers &nbsp;·&nbsp; Generated {now}</div>
</div>
<div class="summary-bar">
  <div class="stat-card"><span class="stat-val" style="color:#60a5fa">{total}</span><span class="stat-label">Total Calls</span></div>
  <div class="stat-card"><span class="stat-val" style="color:#22c55e">{with_data}</span><span class="stat-label">With Data</span></div>
  <div class="stat-card"><span class="stat-val" style="color:#94a3b8">{no_data}</span><span class="stat-label">No Data</span></div>
  <div class="stat-card"><span class="stat-val" style="color:#a5b4fc">{booked}</span><span class="stat-label">Booked</span></div>
  <div class="stat-card"><span class="stat-val" style="color:#f97316">{total_pa}</span><span class="stat-label">Pending Actions</span></div>
</div>
<div class="calls-container">
{cards_html}
</div>
</body>
</html>'''
    return html


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def main():
    print("CSR Test — 10 Calls")
    print("=" * 60)

    results = []

    for i, entry in enumerate(CSR_CALLS):
        num = entry["num"]
        audio_url = entry["audio_url"]
        call_id = entry["call_id"]

        print(f"\n[{num}/10] Audio URL #{num}")

        # If call_id is unknown, try to find it
        if not call_id:
            print(f"  Searching for call by audio URL...")
            call_id = find_call_by_audio_url(audio_url)
            if call_id:
                print(f"  Found call_id: {call_id}")
            else:
                print(f"  Not found in API — skipping")
                results.append({"num": num, "audio_url": audio_url, "call_id": "not_found", "data": None, "status": "not_found"})
                continue

        print(f"  call_id: {call_id}")

        # Fetch summary and detail
        print(f"  Fetching summary...")
        summary = fetch_summary(call_id)
        print(f"  Fetching detail...")
        detail = fetch_detail(call_id)

        if summary:
            print(f"  Summary: OK")
        else:
            print(f"  Summary: NO DATA")

        data = extract_row_data(entry, summary, detail) if summary else None
        results.append({"num": num, "audio_url": audio_url, "call_id": call_id, "data": data, "status": "completed" if data else "no_data"})

    print(f"\n{'='*60}")
    print(f"Updating Excel sheet...")
    update_excel(results)

    print(f"Generating HTML report...")
    html = build_html(results)
    with open(HTML_OUTPUT, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"HTML saved: {HTML_OUTPUT}")

    with_data = sum(1 for r in results if r["data"])
    print(f"\nSUMMARY")
    print(f"  Calls with data: {with_data}/10")
    print(f"  Excel: {EXCEL_PATH}")
    print(f"  HTML:  {HTML_OUTPUT}")

if __name__ == "__main__":
    main()
